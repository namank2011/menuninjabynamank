import json
import datetime
import shutil
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def get_excel_headers(ws) -> List[str]:
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val or "").strip())
    return headers

def join_hash(values: List[Any], fallback_values: List[Any] = None) -> str:
    res = []
    for i, v in enumerate(values):
        val = str(v).strip()
        if val == "" and fallback_values and i < len(fallback_values):
            val = str(fallback_values[i]).strip()
        res.append(val)
    return "#".join(res)

def export_approved_menu(
    template_path: Path,
    output_path: Path,
    items: List[Dict[str, Any]],
    business_name: str
) -> Path:
    # Ensure directories exist
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Copy template to target output file to preserve sheets, styles, formulas and dropdown verification
    shutil.copyfile(str(template_path), str(output_path))
    
    wb = load_workbook(str(output_path))
    if "Product Bulk" not in wb.sheetnames:
        raise ValueError("Missing 'Product Bulk' sheet in the official template.")
        
    ws = wb["Product Bulk"]
    headers = get_excel_headers(ws)
    
    # Locate headers index (1-based)
    header_indices = {name: idx + 1 for idx, name in enumerate(headers) if name}
    
    # Store Row 2 styles/formatting as templates for added rows
    row_style_templates = {}
    if ws.max_row >= 2:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            row_style_templates[col] = {
                "font": cell.font,
                "border": cell.border,
                "fill": cell.fill,
                "number_format": cell.number_format,
                "alignment": cell.alignment,
                "protection": cell.protection
            }
            
        # Delete example rows starting from Row 2
        ws.delete_rows(2, ws.max_row - 1)
        
    approved_items = [it for it in items if it.get("approved")]
    
    for row_idx, item in enumerate(approved_items, start=2):
        # Format variations
        variations = item.get("variations", [])
        var_names = [v.get("name", "") for v in variations]
        sp_prices = [v.get("sellingPrice", "") for v in variations]
        lp_prices = [v.get("listingPrice", "") if v.get("listingPrice") is not None else v.get("sellingPrice", "") for v in variations]
        
        # Format values
        var_str = ""
        sp_str = ""
        lp_str = ""
        
        if len(variations) > 1 or (len(variations) == 1 and var_names[0] != ""):
            # Multi-variation output: join with '#'
            var_str = join_hash(var_names)
            sp_str = join_hash(sp_prices)
            lp_str = join_hash(lp_prices, fallback_values=sp_prices)
        elif len(variations) == 1:
            # Single variation output: direct values
            var_str = ""
            sp_str = str(sp_prices[0])
            lp_str = str(lp_prices[0])
            
        # Determine Item Type and Station dynamically matching user rules
        import re
        liquor_keywords = {
            "beer", "wine", "spirit", "whiskey", "rum", "vodka", "gin", "tequila", "liquor", 
            "shot", "cocktail", "cocktails", "beers", "wines", "liqueur", "alcohol", 
            "champagne", "brandy", "scotch", "cider"
        }
        
        beverage_keywords = {
            "beverage", "drink", "tea", "coffee", "juice", "shake", "mocktail", 
            "soda", "water", "cold", "hot", "milkshake", "lassi", "cooler", "smoothie", 
            "brew", "coke", "pepsi", "sprite", "fanta", "limca", "tonic", "beverages",
            "drinks", "juices", "shakes", "mocktails", "lassis"
        }
        
        cat_lower = str(item.get("categoryName", "")).lower()
        name_lower = str(item.get("productName", "")).lower()
        desc_lower = str(item.get("description", "")).lower()
        
        cat_words = set(re.findall(r'[a-z]+', cat_lower))
        name_words = set(re.findall(r'[a-z]+', name_lower))
        
        # 1. Classify Item Type and Station
        if (cat_words & liquor_keywords) or (name_words & liquor_keywords):
            item_type = "liquor"
            station = "beverage"
        elif (cat_words & beverage_keywords) or (name_words & beverage_keywords):
            item_type = "beverage"
            station = "beverage"
        else:
            item_type = "food"
            station = "kitchen"
            
        # 2. Dynamic Tag Inference (Veg / Non Veg / Egg prediction fallback)
        dietary_tag = str(item.get("dietaryTag", "")).lower().strip()
        if "non" in dietary_tag:
            dietary_tag = "non veg"
        elif "egg" in dietary_tag:
            dietary_tag = "egg"
        elif "veg" in dietary_tag:
            dietary_tag = "veg"
        else:
            # Fallback keyword logic
            non_veg_patterns = [
                r'\bchicken\b', r'\bmutton\b', r'\bfish\b', r'\bprawns?\b', r'\bcrabs?\b', r'\bbeef\b', 
                r'\bpork\b', r'\bbacon\b', r'\bham\b', r'\bsalami\b', r'\bpepperoni\b', r'\bmeat\b', 
                r'\bduck\b', r'\blamb\b', r'\bseafood\b', r'\bturkey\b', r'\blobsters?\b', r'\bshrimps?\b'
            ]
            egg_patterns = [r'\begg\b', r'\bomelette\b', r'\bscrambled\b']
            
            if any(re.search(pattern, name_lower) or re.search(pattern, desc_lower) for pattern in non_veg_patterns):
                dietary_tag = "non veg"
            elif any(re.search(pattern, name_lower) or re.search(pattern, desc_lower) for pattern in egg_patterns):
                dietary_tag = "egg"
            else:
                dietary_tag = "veg"

        # Prepare cell values dictionary matching template header columns
        cell_values = {
            "Category Name*": item.get("categoryName", ""),
            "Product Name*": item.get("productName", ""),
            "Variant Group Name": item.get("variantGroupName", ""),
            "Variation": var_str,
            "Selling Price*": sp_str,
            "Listing Price": lp_str,
            "Master Status*": item.get("masterStatus", "Active"),
            "Menu Status*": item.get("menuStatus", "Active"),
            "Stock Status*": item.get("stockStatus", "Active"),
            "Description": item.get("description", ""),
            "Item Type": item_type,
            "Dietary Tag": dietary_tag,
            "Tax Category": item.get("taxCategory", "Services"),
            "Tax Type": item.get("taxType", "GST"),
            "Tax Value": item.get("taxValue", 5),
            "Item Code": item.get("itemCode", ""),
            "Station": station,
            "Preparation Time": item.get("preparationTime", ""),
            "Image URL 1": item.get("imageUrl1", ""),
            "Image URL 2": item.get("imageUrl2", ""),
            "Image URL 3": item.get("imageUrl3", ""),
            "Get Weight from Weighing Scale": "No"
        }
        
        # Write to cells
        for header_name, col_idx in header_indices.items():
            if header_name in cell_values:
                val = cell_values[header_name]
                ws.cell(row=row_idx, column=col_idx, value=val)
                
        # Copy formatting to new cell
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            style = row_style_templates.get(col_idx)
            if style:
                from copy import copy
                if style["font"]: cell.font = copy(style["font"])
                if style["border"]: cell.border = copy(style["border"])
                if style["fill"]: cell.fill = copy(style["fill"])
                if style["number_format"] is not None: cell.number_format = style["number_format"]
                if style["alignment"]: cell.alignment = copy(style["alignment"])
                if style["protection"]: cell.protection = copy(style["protection"])

    # Adjust width for dynamic categories & names
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header = str(ws.cell(row=1, column=col_idx).value or "")
        if header in {"Category Name*", "Product Name*", "Description"}:
            ws.column_dimensions[letter].width = 28 if header != "Description" else 42
        elif header in {"Variation", "Selling Price*", "Listing Price"}:
            ws.column_dimensions[letter].width = 18

    wb.save(str(output_path))
    wb.close()
    return output_path

def generate_review_report(
    draft_meta: Dict[str, Any],
    audit_logs: List[Dict[str, Any]],
    report_json_path: Path,
    report_txt_path: Path
):
    items = draft_meta.get("items", [])
    files = draft_meta.get("files", [])
    
    total_detected = len(items)
    approved_items = [it for it in items if it.get("approved")]
    excluded_items = [it for it in items if not it.get("approved")]
    
    categories = sorted(list(set(it.get("categoryName", "Uncategorized") for it in items)))
    var_items_count = sum(1 for it in items if len(it.get("variations", [])) > 1 or (len(it.get("variations", [])) == 1 and it.get("variations", [{}])[0].get("name") != ""))
    non_var_items_count = total_detected - var_items_count
    
    # Audit trail summaries
    price_corrections = sum(1 for log in audit_logs if "sellingPrice" in log.get("details", "") or "price" in log.get("details", "").lower())
    cat_corrections = sum(1 for log in audit_logs if "category" in log.get("details", "").lower())
    var_corrections = sum(1 for log in audit_logs if "variation" in log.get("details", "").lower() or "variant" in log.get("details", "").lower())
    
    report_data = {
        "businessName": draft_meta.get("businessName", "Default Business"),
        "timestamp": datetime.datetime.now().isoformat(),
        "finalApprovedBy": "Human Reviewer",
        "sourceFiles": [f.get("name") for f in files],
        "metrics": {
            "totalProductsDetected": total_detected,
            "totalApproved": len(approved_items),
            "totalExcluded": len(excluded_items),
            "categoriesCreatedCount": len(categories),
            "productsWithVariations": var_items_count,
            "productsWithoutVariations": non_var_items_count
        },
        "exclusions": [
            {
                "productName": it.get("productName"),
                "category": it.get("categoryName"),
                "reason": "Not approved by reviewer during human feedback stage."
            } for it in excluded_items
        ],
        "corrections": {
            "pricesCorrected": price_corrections,
            "categoriesCorrected": cat_corrections,
            "variationMappingsCorrected": var_corrections
        },
        "auditLogsSummary": audit_logs[:50]  # List top 50 logs
    }
    
    # Write JSON report
    report_json_path.parent.mkdir(parents=True, exist_ok=True)
    report_json_path.write_text(json.dumps(report_data, indent=2, ensure_ascii=False), encoding="utf-8")
    
    # Write Text report
    txt_report = f"""========================================================================
                      SHOPVERSE REVIEW REPORT
========================================================================
Business Name: {report_data["businessName"]}
Report Date: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Final Approver: {report_data["finalApprovedBy"]}

Source Files:
{newline_join([f" - {name}" for name in report_data["sourceFiles"]])}

------------------------------------------------------------------------
METRICS SUMMARY
------------------------------------------------------------------------
* Total Products Processed: {total_detected}
* Approved Products (Exported): {len(approved_items)}
* Excluded Products: {len(excluded_items)}
* Categories Exported: {len(categories)} ({", ".join(categories)})
* Products with Variations: {var_items_count}
* Products without Variations: {non_var_items_count}

------------------------------------------------------------------------
USER CORRECTION COUNTS (AUDITED)
------------------------------------------------------------------------
* Category re-mappings: {cat_corrections}
* Variation sequence corrections: {var_corrections}
* Pricing adjustments: {price_corrections}

------------------------------------------------------------------------
EXCLUDED PRODUCTS LIST
------------------------------------------------------------------------
{newline_join([f" - {it['productName']} ({it['category']}): {it['reason']}" for it in report_data["exclusions"]]) if report_data["exclusions"] else "None (All products approved for bulk upload)"}

------------------------------------------------------------------------
AUDIT LOG (LAST 20 LOGS)
------------------------------------------------------------------------
{newline_join([f" [{log['timestamp'][:19]}] {log['user']}: {log['details']}" for log in report_data["auditLogsSummary"][:20]])}

========================================================================
"""
    report_txt_path.write_text(txt_report, encoding="utf-8")

def newline_join(lst: List[str]) -> str:
    return "\n".join(lst)
