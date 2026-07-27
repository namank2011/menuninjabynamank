from __future__ import annotations

import json
import re
from copy import copy
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path
backend_dir = Path(__file__).resolve().parent
if str(backend_dir) not in sys.path:
    sys.path.insert(0, str(backend_dir))

from schemas import AddOnGroup, MenuExtraction, MenuItem

PRODUCT_SHEET = "Product Bulk"
MANDATORY_HEADERS = [
    "Category Name*",
    "Product Name*",
    "Selling Price*",
    "Master Status*",
    "Menu Status*",
    "Stock Status*",
]

VALID_DIETARY = {"Veg", "Non-Veg", "Egg", ""}
VALID_STATUS = {"Active", "Inactive"}
VALID_TAX_CATEGORY = {"Goods", "Services", ""}
VALID_TAX_TYPE = {"GST", "VAT", ""}


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _clean_price(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value)
    text = text.replace(",", "")
    match = re.search(r"\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def _format_price(value: Any) -> str:
    price = _clean_price(value)
    if price is None:
        return ""
    if float(price).is_integer():
        return str(int(price))
    return str(round(float(price), 2))


def _join_hash(values: List[str]) -> str:
    return "#".join(str(v).strip() for v in values if str(v).strip() != "")


def _variation_fields(item: MenuItem) -> Tuple[str, str, str]:
    variations = item.variations or []
    names = [_clean_text(v.name) for v in variations]
    prices = [_format_price(v.price) for v in variations]
    listing_prices = [_format_price(v.listing_price if v.listing_price is not None else v.price) for v in variations]

    # One simple price: keep Variation blank and price as single value.
    if len(variations) <= 1 and (not names or names[0] == ""):
        return "", prices[0] if prices else "", listing_prices[0] if listing_prices else ""

    # Variation and prices must be separated by # as required.
    return _join_hash(names), _join_hash(prices), _join_hash(listing_prices)


def _get_header_map(ws) -> Dict[str, int]:
    """Return first occurrence of every header. Duplicate add-on headers are handled separately."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and val not in headers:
            headers[str(val).strip()] = col
    return headers


def _find_duplicate_header_groups(ws, group_headers: List[str]) -> List[Dict[str, int]]:
    """Find repeating add-on groups based on the template headers."""
    groups = []
    cols = [(col, str(ws.cell(row=1, column=col).value or "").strip()) for col in range(1, ws.max_column + 1)]
    start_cols = [col for col, header in cols if header == "Add On Category"]
    for start in start_cols:
        group = {}
        for offset, header in enumerate(group_headers):
            current_col = start + offset
            current_header = str(ws.cell(row=1, column=current_col).value or "").strip()
            if current_header == header:
                group[header] = current_col
        if "Add On Category" in group and "Add On Products" in group:
            groups.append(group)
    return groups


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)
        if src.border:
            dst.border = copy(src.border)


def _write_by_header(ws, header_map: Dict[str, int], row_idx: int, header: str, value: Any) -> None:
    col = header_map.get(header)
    if col:
        ws.cell(row=row_idx, column=col).value = value


def _validate_item(item: MenuItem) -> List[str]:
    warnings = []
    if not _clean_text(item.category):
        warnings.append("Missing category")
    if not _clean_text(item.product_name):
        warnings.append("Missing product name")

    variation, prices, _ = _variation_fields(item)
    if not prices:
        warnings.append("Missing selling price")
    if item.dietary_tag not in VALID_DIETARY:
        warnings.append(f"Invalid dietary tag: {item.dietary_tag}")
    if item.tax_category not in VALID_TAX_CATEGORY:
        warnings.append(f"Invalid tax category: {item.tax_category}")
    if item.tax_type not in VALID_TAX_TYPE:
        warnings.append(f"Invalid tax type: {item.tax_type}")
    if item.confidence is not None and item.confidence < 0.75:
        warnings.append("Low confidence - review item")
    return warnings


def write_bulk_upload_excel(
    template_path: str | Path,
    extraction: MenuExtraction,
    output_path: str | Path,
    review_json_path: str | Path | None = None,
) -> Dict[str, Any]:
    template_path = Path(template_path)
    output_path = Path(output_path)
    wb = load_workbook(str(template_path))

    if PRODUCT_SHEET not in wb.sheetnames:
        raise ValueError(f"Template must contain sheet named '{PRODUCT_SHEET}'. Found: {wb.sheetnames}")

    ws = wb[PRODUCT_SHEET]
    header_map = _get_header_map(ws)

    missing_headers = [h for h in MANDATORY_HEADERS if h not in header_map]
    if missing_headers:
        raise ValueError(f"Missing mandatory headers in template: {missing_headers}")

    # Preserve example row styling, then clear all old/example rows.
    style_source_row = 2 if ws.max_row >= 2 else 1
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    add_on_group_headers = [
        "Add On Category", "Add On Products", "Add On Rule", "Max Addon Qty", "Max Qty",
        "Add On Required", "Min Addon Qty", "Min Qty", "Chargeable Count"
    ]
    add_on_groups = _find_duplicate_header_groups(ws, add_on_group_headers)

    review_rows = []
    written = 0

    for item in extraction.items:
        if not _clean_text(item.product_name):
            continue

        row_idx = 2 + written
        if row_idx != style_source_row:
            _copy_row_style(ws, style_source_row, row_idx)

        variation, selling_price, listing_price = _variation_fields(item)
        warnings = _validate_item(item)

        _write_by_header(ws, header_map, row_idx, "Category Name*", _clean_text(item.category))
        _write_by_header(ws, header_map, row_idx, "Product Name*", _clean_text(item.product_name))
        _write_by_header(ws, header_map, row_idx, "Variant Group Name", "")
        _write_by_header(ws, header_map, row_idx, "Variation", variation)
        _write_by_header(ws, header_map, row_idx, "Selling Price*", selling_price)
        _write_by_header(ws, header_map, row_idx, "Listing Price", listing_price or selling_price)
        _write_by_header(ws, header_map, row_idx, "Master Status*", "Active")
        _write_by_header(ws, header_map, row_idx, "Menu Status*", "Active")
        _write_by_header(ws, header_map, row_idx, "Stock Status*", "Active")
        _write_by_header(ws, header_map, row_idx, "Description", _clean_text(item.description))
        _write_by_header(ws, header_map, row_idx, "Item Type", _clean_text(item.item_type))
        _write_by_header(ws, header_map, row_idx, "Tax Category", item.tax_category or "Services")
        _write_by_header(ws, header_map, row_idx, "Tax Type", item.tax_type or "GST")
        _write_by_header(ws, header_map, row_idx, "Tax Value", item.tax_value if item.tax_value is not None else 5)
        _write_by_header(ws, header_map, row_idx, "Item Code", _clean_text(item.item_code))
        _write_by_header(ws, header_map, row_idx, "Station", _clean_text(item.station or "Kitchen"))
        _write_by_header(ws, header_map, row_idx, "Preparation Time", _clean_text(item.preparation_time))
        _write_by_header(ws, header_map, row_idx, "Dietary Tag", item.dietary_tag if item.dietary_tag in VALID_DIETARY else "")
        _write_by_header(ws, header_map, row_idx, "Get Weight from Weighing Scale", "No")
        _write_by_header(ws, header_map, row_idx, "Image URL 1", _clean_text(item.image_url_1))

        # Optional add-on groups: fill max 3 groups based on template structure.
        for group_index, addon in enumerate(item.add_on_groups[:len(add_on_groups)]):
            group_cols = add_on_groups[group_index]
            ws.cell(row=row_idx, column=group_cols.get("Add On Category")).value = addon.category
            ws.cell(row=row_idx, column=group_cols.get("Add On Products")).value = "#".join(addon.products)
            ws.cell(row=row_idx, column=group_cols.get("Add On Rule")).value = addon.rule
            if "Max Addon Qty" in group_cols:
                ws.cell(row=row_idx, column=group_cols["Max Addon Qty"]).value = addon.max_qty
            if "Max Qty" in group_cols:
                ws.cell(row=row_idx, column=group_cols["Max Qty"]).value = addon.max_qty
            if "Add On Required" in group_cols:
                ws.cell(row=row_idx, column=group_cols["Add On Required"]).value = addon.required
            if "Min Addon Qty" in group_cols:
                ws.cell(row=row_idx, column=group_cols["Min Addon Qty"]).value = addon.min_qty
            if "Min Qty" in group_cols:
                ws.cell(row=row_idx, column=group_cols["Min Qty"]).value = addon.min_qty
            if "Chargeable Count" in group_cols:
                ws.cell(row=row_idx, column=group_cols["Chargeable Count"]).value = addon.chargeable_count

        review_rows.append({
            "row": row_idx,
            "category": item.category,
            "product_name": item.product_name,
            "variation": variation,
            "selling_price": selling_price,
            "confidence": item.confidence,
            "warnings": warnings,
            "source_text": item.source_text,
        })
        written += 1

    # Friendly column widths without destroying template format.
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header = str(ws.cell(row=1, column=col_idx).value or "")
        if header in {"Category Name*", "Product Name*", "Description"}:
            ws.column_dimensions[letter].width = 28 if header != "Description" else 42
        elif header in {"Variation", "Selling Price*", "Listing Price"}:
            ws.column_dimensions[letter].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    result = {
        "output_xlsx": str(output_path),
        "total_items_written": written,
        "review_required_count": sum(1 for r in review_rows if r["warnings"]),
        "review_rows": review_rows,
    }

    if review_json_path:
        review_json_path = Path(review_json_path)
        review_json_path.parent.mkdir(parents=True, exist_ok=True)
        review_json_path.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
        result["review_json"] = str(review_json_path)

    return result
