import re
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import load_workbook

# Hardcoded fallbacks
VALIDATION_DEFAULTS = {
    "Status": ["Active", "Inactive"],
    "Tax Category": ["Goods", "Services"],
    "Tax Type": ["GST", "VAT"],
    "Dietary Tag": ["Veg", "Non-Veg", "Egg", ""],
    "Boolean": ["Yes", "No"],
    "Allergen Type": ["Gluten", "Crustacean", "Egg", "Fish", "Nuts", "Peanut", "Soyabeans", "Milk", "Sulphite"],
    "Portion Size Unit": ["grams", "kg", "inches", "litre", "ml", "ounces", "pounds", "serves", "slices", "cms", "piece", "scoop"],
    "Serving Info": [
        "For 1 to 2 people", "For 2 to 3 people", "For 3 to 4 people", 
        "For 4 to 5 people", "For 5 to 6 people", "For 6 to 7 people", 
        "For 7 to 8 people", "For 8 to 9 people", "For 9 to 10 people", 
        "For more than 10 people"
    ],
    "Add On Rule": ["Select Only One", "Select Multiple", "Select Custom"],
    "Incentive Type": ["AMT", "%"]
}

def load_validation_lists(template_path: Path) -> Dict[str, List[str]]:
    lists = {}
    try:
        wb = load_workbook(str(template_path), read_only=True)
        if "Validation" in wb.sheetnames:
            ws = wb["Validation"]
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_str = str(header).strip()
                    values = []
                    for row in range(2, 60):
                        val = ws.cell(row=row, column=col).value
                        if val is not None:
                            values.append(str(val).strip())
                        else:
                            break
                    lists[header_str] = values
        wb.close()
    except Exception:
        pass
    
    # Merge with default fallbacks for missing lists
    for k, v in VALIDATION_DEFAULTS.items():
        if k not in lists or not lists[k]:
            lists[k] = v
            
    # Normalize empty tags
    if "" not in lists.get("Dietary Tag", []):
        lists.setdefault("Dietary Tag", []).append("")
        
    return lists

def validate_item(item: Dict[str, Any], validation_lists: Dict[str, List[str]], all_items: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    errors = []
    
    # 1. Product Name*
    product_name = str(item.get("productName", "")).strip()
    if not product_name:
        errors.append({"type": "Blocking Error", "field": "productName", "message": "Product Name is required."})
    elif any(placeholder in product_name.lower() for placeholder in ["[insert", "ocr error", "unknown", "placeholder"]):
        errors.append({"type": "Warning", "field": "productName", "message": "Product name contains potential OCR placeholder text."})

    # 2. Category Name*
    cat_name = str(item.get("categoryName", "")).strip()
    if not cat_name or cat_name.lower() in ["uncategorized", "category review required"]:
        errors.append({"type": "Warning", "field": "categoryName", "message": "Product is uncategorized. Needs category mapping."})
        
    # 3. Status checks
    for status_field in ["masterStatus", "menuStatus", "stockStatus"]:
        val = item.get(status_field, "Active")
        if val not in validation_lists.get("Status", VALIDATION_DEFAULTS["Status"]):
            errors.append({"type": "Blocking Error", "field": status_field, "message": f"Invalid {status_field}: '{val}'. Allowed values: {validation_lists.get('Status')}"})
            
    # 4. Dietary Tag
    diet_tag = item.get("dietaryTag", "")
    allowed_diet_tags = validation_lists.get("Dietary Tag", VALIDATION_DEFAULTS["Dietary Tag"])
    allowed_diet_tags_lower = [str(t).lower().strip() for t in allowed_diet_tags]
    # Explicitly support 'veg', 'non veg', 'egg', 'non-veg'
    for x in ["veg", "non veg", "egg", "non-veg"]:
        if x not in allowed_diet_tags_lower:
            allowed_diet_tags_lower.append(x)
            
    if str(diet_tag).lower().strip() not in allowed_diet_tags_lower:
        errors.append({"type": "Blocking Error", "field": "dietaryTag", "message": f"Invalid Dietary Tag: '{diet_tag}'."})
        
    # 5. Tax validation
    tax_cat = item.get("taxCategory", "Services")
    if tax_cat not in validation_lists.get("Tax Category", VALIDATION_DEFAULTS["Tax Category"]):
        errors.append({"type": "Blocking Error", "field": "taxCategory", "message": f"Invalid Tax Category: '{tax_cat}'."})
        
    tax_type = item.get("taxType", "GST")
    if tax_type not in validation_lists.get("TaxType", VALIDATION_DEFAULTS["Tax Type"]):
        errors.append({"type": "Blocking Error", "field": "taxType", "message": f"Invalid Tax Type: '{tax_type}'."})
        
    tax_val = item.get("taxValue")
    if tax_val is not None:
        try:
            f_val = float(tax_val)
            if f_val < 0:
                errors.append({"type": "Blocking Error", "field": "taxValue", "message": "Tax Value cannot be negative."})
        except ValueError:
            errors.append({"type": "Blocking Error", "field": "taxValue", "message": "Tax Value must be numeric."})
            
    # 6. Variations & Prices (Crucial validation logic)
    variations = item.get("variations", [])
    variant_group = str(item.get("variantGroupName", "")).strip()
    
    if len(variations) == 0:
        errors.append({"type": "Blocking Error", "field": "variations", "message": "At least one price/variation is required."})
    elif len(variations) == 1 and not variations[0].get("name"):
        # Single price item
        price = variations[0].get("sellingPrice")
        if price is None or str(price).strip() == "":
            errors.append({"type": "Blocking Error", "field": "sellingPrice", "message": "Selling Price is required."})
        else:
            try:
                f_price = float(price)
                if f_price < 0:
                    errors.append({"type": "Blocking Error", "field": "sellingPrice", "message": "Price cannot be negative."})
            except ValueError:
                errors.append({"type": "Blocking Error", "field": "sellingPrice", "message": "Price must be numeric."})
                
        list_price = variations[0].get("listingPrice")
        if list_price is not None and str(list_price).strip() != "":
            try:
                f_lp = float(list_price)
                f_sp = float(price or 0)
                if f_lp < f_sp:
                    errors.append({"type": "Blocking Error", "field": "listingPrice", "message": "Listing Price (MRP) cannot be less than Selling Price."})
            except ValueError:
                 errors.append({"type": "Blocking Error", "field": "listingPrice", "message": "Listing Price must be numeric."})
    else:
        # Multi-variation item
        if not variant_group:
            errors.append({"type": "Warning", "field": "variantGroupName", "message": "Multiple variations exist but Variant Group Name is blank."})
            
        for i, var in enumerate(variations):
            var_name = str(var.get("name", "")).strip()
            if not var_name:
                errors.append({"type": "Blocking Error", "field": f"variations[{i}].name", "message": f"Variation {i+1} name cannot be empty."})
            if "#" in var_name:
                errors.append({"type": "Blocking Error", "field": f"variations[{i}].name", "message": "Variation names cannot contain '#' character."})
                
            sp = var.get("sellingPrice")
            if sp is None or str(sp).strip() == "":
                errors.append({"type": "Blocking Error", "field": f"variations[{i}].sellingPrice", "message": f"Selling price for variation '{var_name}' is required."})
            else:
                try:
                    f_sp = float(sp)
                    if f_sp < 0:
                        errors.append({"type": "Blocking Error", "field": f"variations[{i}].sellingPrice", "message": "Price cannot be negative."})
                except ValueError:
                    errors.append({"type": "Blocking Error", "field": f"variations[{i}].sellingPrice", "message": "Price must be numeric."})
                    
            lp = var.get("listingPrice")
            if lp is not None and str(lp).strip() != "":
                try:
                    f_lp = float(lp)
                    f_sp = float(sp or 0)
                    if f_lp < f_sp:
                        errors.append({"type": "Blocking Error", "field": f"variations[{i}].listingPrice", "message": "Listing Price (MRP) cannot be less than Selling Price."})
                except ValueError:
                    errors.append({"type": "Blocking Error", "field": f"variations[{i}].listingPrice", "message": "Listing Price must be numeric."})

    # 7. Low confidence score warnings
    confidence = item.get("confidence")
    if confidence is not None:
        try:
            f_conf = float(confidence)
            if f_conf < 0.75:
                errors.append({"type": "Warning", "field": "confidence", "message": f"Low extraction confidence ({int(f_conf * 100)}%). Needs validation."})
        except ValueError:
            pass

    # 8. Duplicate detection (Section 21)
    norm_name = product_name.lower().replace(" ", "")
    norm_cat = cat_name.lower().replace(" ", "")
    
    is_duplicate = False
    for other in all_items:
        if other.get("id") == item.get("id"):
            continue
        other_name = str(other.get("productName", "")).strip().lower().replace(" ", "")
        other_cat = str(other.get("categoryName", "")).strip().lower().replace(" ", "")
        if other_name == norm_name and other_cat == norm_cat:
            is_duplicate = True
            break
            
    if is_duplicate:
        errors.append({"type": "Warning", "field": "productName", "message": "Possible duplicate product detected in the same category."})

    return errors

def validate_menu(items: List[Dict[str, Any]], template_path: Path) -> List[Dict[str, Any]]:
    validation_lists = load_validation_lists(template_path)
    validated_items = []
    
    for item in items:
        # Clone item and run validation
        item_copy = dict(item)
        errs = validate_item(item_copy, validation_lists, items)
        item_copy["validationErrors"] = errs
        
        # Set status based on validation errors
        if any(e["type"] == "Blocking Error" for e in errs):
            item_copy["reviewStatus"] = "Blocked"
        elif any(e["type"] == "Warning" for e in errs):
            item_copy["reviewStatus"] = "Review Required"
        else:
            # Keep original review status unless it needs review
            if item_copy.get("reviewStatus") in ["Blocked", "Review Required"]:
                item_copy["reviewStatus"] = "Reviewed"
                
        validated_items.append(item_copy)
        
    return validated_items
