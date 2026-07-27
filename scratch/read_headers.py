from openpyxl import load_workbook
from pathlib import Path

template_path = Path("c:/Users/Amonex/QuickMenuAgent/backend/templates/Bulk_Upload_Sheet_Format.xlsx")
wb = load_workbook(str(template_path))
ws = wb["Product Bulk"]
headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
print("Product Bulk Headers:", [h for h in headers if h])
if "Validation" in wb.sheetnames:
    ws_val = wb["Validation"]
    print("Validation Sheets:")
    for col in range(1, ws_val.max_column + 1):
        h = ws_val.cell(row=1, column=col).value
        if h:
            vals = []
            for r in range(2, 20):
                v = ws_val.cell(row=r, column=col).value
                if v is not None:
                    vals.append(v)
            print(f"  {h}: {vals}")
wb.close()
