import os
import requests
import json
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

# Define constants
BASE_URL = "http://127.0.0.1:8000/api/drafts"
IMAGE_FILE = Path(r"c:\Users\Amonex\QuickMenuAgent\uploads\aedb656387_WhatsApp Image 2026-07-27 at 1.48.13 PM.jpeg")
GEMINI_KEY = os.getenv("GEMINI_API_KEY", "")

def verify_flow():
    if not IMAGE_FILE.exists():
        print(f"Error: Sample image not found at {IMAGE_FILE}")
        return

    # Use a requests session to automatically persist backend auth cookie (session_token)
    session = requests.Session()

    # Log in first
    print("Logging in to obtain session token...")
    login_url = "http://127.0.0.1:8000/api/auth/login"
    login_res = session.post(login_url, json={
        "email": "namankshetri2@gmail.com",
        "password": "2011@Naman"
    })
    if login_res.status_code != 200:
        print(f"Authentication failed (status code: {login_res.status_code}): {login_res.text}")
        return
    print("Authenticated successfully.")
    
    session_token = session.cookies.get("session_token") or login_res.cookies.get("session_token")
    cookie_header = {"Cookie": f"session_token={session_token}"} if session_token else {}
    print(f"Session token obtained: {session_token[:20] if session_token else 'None'}...")

    print("--- 1. Testing Regular Review & Export Flow ---")
    headers = {"X-Gemini-API-Key": GEMINI_KEY, **cookie_header}
    
    # Create draft
    files = [("menu_files", (IMAGE_FILE.name, open(IMAGE_FILE, "rb"), "image/jpeg"))]
    data = {
        "business_name": "Super Bistro",
        "default_template": "true",
        "extraction_engine": "gemini",
        "tax_category": "Services",
        "tax_type": "GST",
        "tax_value": "5.0",
        "master_status": "Active",
        "menu_status": "Active",
        "stock_status": "Active",
        "station": "Kitchen",
        "preparation_time": "15 mins",
        "default_dietary": "",
        "direct_approve": "false"  # Manual review flow
    }
    
    res_post = session.post(BASE_URL, headers=headers, files=files, data=data, timeout=120)
    print(f"Post draft status: {res_post.status_code}")
    post_json = res_post.json()
    draft_id = post_json.get("draftId")
    print(f"Created Draft ID: {draft_id}")
    
    if not draft_id:
        print("Failed to create draft.")
        return
        
    # Get draft details
    res_get = session.get(f"{BASE_URL}/{draft_id}", headers=headers)
    draft_details = res_get.json()
    print("Draft items count:", len(draft_details.get("items", [])))
    
    # Mark items approved for export
    print("Simulating human review: Approving all items...")
    conn = sqlite3.connect("outputs/shopverse_agent.db")
    cursor = conn.cursor()
    cursor.execute("UPDATE draft_items SET approved = 1, review_status = 'Reviewed' WHERE draft_id = ?", (draft_id,))
    conn.commit()
    conn.close()
    
    # Approve and export draft
    print("Calling approval / export endpoint...")
    res_approve = session.post(f"{BASE_URL}/{draft_id}/approve", headers=headers, json={"approvedAgreement": True})
    approve_json = res_approve.json()
    print("Approve API Response Keys:", list(approve_json.keys()))
    print("Expected Excel Filename:", approve_json.get("outputFile"))
    print("Expected JSON Report Filename:", approve_json.get("reviewReportJson"))
    print("Expected TXT Report Filename:", approve_json.get("reviewReportTxt"))
    
    # Verify Excel filename has the business name in it
    output_filename = approve_json.get("outputFile")
    assert "Super_Bistro" in output_filename, f"Business name not in output filename: {output_filename}"
    print("SUCCESS: Filename matches business name on manual review export!")
    
    print("\n--- 2. Testing One-Click Quick Bulk Export Flow ---")
    data["direct_approve"] = "true"
    data["business_name"] = "Bar Grill House"
    files = [("menu_files", (IMAGE_FILE.name, open(IMAGE_FILE, "rb"), "image/jpeg"))]
    
    res_direct = session.post(BASE_URL, headers=headers, files=files, data=data, timeout=120)
    print(f"Direct export status: {res_direct.status_code}")
    direct_json = res_direct.json()
    direct_xlsx = direct_json.get("outputFile")
    print("Direct Export Excel Name:", direct_xlsx)
    
    assert "Bar_Grill_House" in direct_xlsx, f"Business name not in direct output filename: {direct_xlsx}"
    print("SUCCESS: Filename matches business name on direct export!")
    
    # Inspect physical exported file to verify Station classification
    out_dir = Path("outputs")
    xlsx_path = out_dir / direct_xlsx
    if xlsx_path.exists():
        print(f"Found exported file at: {xlsx_path}")
        wb = load_workbook(str(xlsx_path))
        ws = wb["Product Bulk"]
        
        # Read headers
        headers_row = [cell.value for cell in ws[1]]
        station_idx = headers_row.index("Station") + 1
        cat_idx = headers_row.index("Category Name*") + 1
        name_idx = headers_row.index("Product Name*") + 1
        
        print("\nVerifying written stations in the Excel file sheets:")
        for r in range(2, ws.max_row + 1):
            prod_name = ws.cell(r, name_idx).value
            cat_name = ws.cell(r, cat_idx).value
            station = ws.cell(r, station_idx).value
            print(f"Product: {prod_name} | Category: {cat_name} | Station: {station}")
            assert station in ["Kitchen", "Bar"], f"Invalid station: {station}"
            
        wb.close()
        print("SUCCESS: Excel file verified. Station is capitalized Kitchen/Bar and properly mapped!")
    else:
        print(f"Error: Direct export file was not saved at {xlsx_path}")

if __name__ == "__main__":
    verify_flow()
